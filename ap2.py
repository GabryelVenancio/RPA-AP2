import sqlite3
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime

TRADUCAO_PAISES = {
    'espanha': 'spain',
    'noruega': 'norway',
    'suíça': 'switzerland',
    'brasil': 'brazil',
    'eua': 'usa',
    'estados unidos': 'usa',
    'japão': 'japan'
}

def traduzir_pais(nome_pt):
    """Traduz nome do país em português para inglês"""
    nome_pt = nome_pt.lower().strip()
    return TRADUCAO_PAISES.get(nome_pt, nome_pt)

def parte_1_extracao_paises():
    paises = []
    for i in range(3):
        pais = input(f"Digite o nome do {i+1}º país: ").strip()
        paises.append(pais)
    
    conn = sqlite3.connect('paises.db')
    cursor = conn.cursor()
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS paises (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome_comum TEXT,
        nome_oficial TEXT,
        capital TEXT,
        continente TEXT,
        regiao TEXT,
        subregiao TEXT,
        populacao INTEGER,
        area REAL,
        moeda_nome TEXT,
        moeda_simbolo TEXT,
        idioma_principal TEXT,
        fuso_horario TEXT,
        url_bandeira TEXT
    )
    ''')
    conn.commit()
    
    for pais in paises:
        try:
            pais_ingles = traduzir_pais(pais)
            response = requests.get(f'https://restcountries.com/v3.1/name/{pais_ingles}')
            response.raise_for_status()
            dados = response.json()[0]
            
            moedas = dados.get('currencies', {})
            moeda_nome = moeda_simbolo = None
            if moedas:
                primeira_moeda = list(moedas.values())[0]
                moeda_nome = primeira_moeda.get('name', '')
                moeda_simbolo = primeira_moeda.get('symbol', '')
            
            idiomas = dados.get('languages', {})
            idioma_principal = list(idiomas.values())[0] if idiomas else None
            
            cursor.execute('''
            INSERT INTO paises (
                nome_comum, nome_oficial, capital, continente, regiao, subregiao,
                populacao, area, moeda_nome, moeda_simbolo, idioma_principal,
                fuso_horario, url_bandeira
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                dados.get('name', {}).get('common', ''),
                dados.get('name', {}).get('official', ''),
                ', '.join(dados.get('capital', [])),
                dados.get('continent', ''),
                dados.get('region', ''),
                dados.get('subregion', ''),
                dados.get('population', 0),
                dados.get('area', 0),
                moeda_nome,
                moeda_simbolo,
                idioma_principal,
                ', '.join(dados.get('timezones', [])),
                dados.get('flags', {}).get('png', '')
            ))
            conn.commit()
            print(f"Dados de {pais} inseridos com sucesso!")
            
        except Exception as e:
            print(f"Erro ao processar {pais}: {str(e)}")
    
    conn.close()

def parte_2_web_scraping_livros():
    conn = sqlite3.connect('livraria.db')
    cursor = conn.cursor()
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS livros (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titulo TEXT,
        preco REAL,
        avaliacao INTEGER,
        disponibilidade TEXT
    )
    ''')
    conn.commit()
    
    try:
        url = 'https://books.toscrape.com/'
        response = requests.get(url)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'html.parser')
        livros = soup.find_all('article', class_='product_pod')[:10]
        
        for livro in livros:
            titulo = livro.h3.a['title']

            preco_texto = livro.find('p', class_='price_color').text
            preco = float(preco_texto.replace('Â', '').replace('£', ''))
            
            rating_classes = {
                'One': 1,
                'Two': 2,
                'Three': 3,
                'Four': 4,
                'Five': 5
            }
            rating_class = livro.p['class'][1]
            avaliacao = rating_classes.get(rating_class, 0)
            
            disponibilidade = livro.find('p', class_='instock availability').text.strip()
            
            cursor.execute('''
            INSERT INTO livros (titulo, preco, avaliacao, disponibilidade)
            VALUES (?, ?, ?, ?)
            ''', (titulo, preco, avaliacao, disponibilidade))
        
        conn.commit()
        print("Dados dos livros inseridos com sucesso!")
        
    except Exception as e:
        print(f"Erro ao fazer scraping: {str(e)}")
    finally:
        conn.close()

def parte_3_gerar_relatorio():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        ws['A1'] = "Relatório de Dados Extraídos"
        ws['A2'] = f"Gerado por: Gabryel"
        ws['A3'] = f"Data de geração: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"

        ws.cell(row=5, column=1, value="Dados dos Países")
        headers_paises = [
            "Nome Comum", "Nome Oficial", "Capital", "Continente", 
            "Região", "Sub-região", "População", "Área", 
            "Moeda (Nome)", "Moeda (Símbolo)", "Idioma Principal", 
            "Fuso Horário", "URL da Bandeira"
        ]
        
        for col, header in enumerate(headers_paises, start=1):
            ws.cell(row=6, column=col, value=header)
        
        conn = sqlite3.connect('paises.db')
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM paises')
        paises = cursor.fetchall()
        
        for row_idx, pais in enumerate(paises, start=7):
            for col_idx, value in enumerate(pais[1:], start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        inicio_livros = 7 + len(paises) + 3
        ws.cell(row=inicio_livros, column=1, value="Dados dos Livros")
        headers_livros = ["Título", "Preço", "Avaliação", "Disponibilidade"]
        
        for col, header in enumerate(headers_livros, start=1):
            ws.cell(row=inicio_livros+1, column=col, value=header)
        
        conn = sqlite3.connect('livraria.db')
        cursor = conn.cursor()
        cursor.execute('SELECT titulo, preco, avaliacao, disponibilidade FROM livros')
        livros = cursor.fetchall()
        
        for row_idx, livro in enumerate(livros, start=inicio_livros+2):
            for col_idx, value in enumerate(livro, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        nome_arquivo = f"relatorio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(nome_arquivo)
        print(f"Relatório gerado com sucesso: {nome_arquivo}")
        
    except Exception as e:
        print(f"Erro ao gerar relatório: {str(e)}")
    finally:
        conn.close()

def main():
    print("=== Sistema de Extração de Dados ===")
    print("\nParte 1: Extração de dados de países")
    parte_1_extracao_paises()
    
    print("\nParte 2: Web scraping de livros")
    parte_2_web_scraping_livros()
    
    print("\nParte 3: Gerando relatório")
    parte_3_gerar_relatorio()
    
    print("\nProcesso concluído com sucesso!")

if __name__ == "__main__":
    main()